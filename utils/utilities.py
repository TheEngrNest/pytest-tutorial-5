import openpyxl


def get_testdata(sheet_name, testdata_file_name='test_data/calculator_test_data.xlsx'):
    wb_obj = openpyxl.load_workbook(testdata_file_name)
    data_sheet = wb_obj[sheet_name]
    max_rw = data_sheet.max_row
    max_cl = data_sheet.max_column
    test_data = list()
    for row in data_sheet.iter_rows(min_row=2, max_col=max_cl, max_row=max_rw, values_only=True):
        test_data.append(row)
        print(row)
    return test_data
